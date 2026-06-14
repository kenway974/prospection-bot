"""
Interface Streamlit — Prospection B2B automatisée.
Lance avec : streamlit run app.py
"""

import json
import os
import queue
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from functools import partial
from typing import Optional

import streamlit as st

# ---------------------------------------------------------------------------
# Config page (doit être le 1er appel Streamlit)
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Prospection B2B",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# CSS custom
# ---------------------------------------------------------------------------
st.markdown("""
<style>
    .main { background-color: #0f1117; }

    /* Bouton */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white; border: none; border-radius: 8px;
        padding: 0.6rem 2rem; font-size: 1rem; font-weight: 600;
        width: 100%; transition: opacity 0.2s;
    }
    .stButton > button:hover { opacity: 0.85; }

    /* Metric cards */
    .metric-card {
        background: #1e2130; border-radius: 10px;
        padding: 1rem; text-align: center; border: 1px solid #2d3250;
        margin-bottom: 8px;
    }
    .metric-value { font-size: 1.6rem; font-weight: 700; color: #667eea; }
    .metric-label { font-size: 0.75rem; color: #888; margin-top: 4px; }

    /* Log box */
    .log-box {
        background: #0d1117; border: 1px solid #2d3250;
        border-radius: 8px; padding: 1rem;
        font-family: monospace; font-size: 0.75rem;
        height: 250px; overflow-y: auto; color: #cdd6f4;
    }

    /* Issue chips */
    .issue-chip {
        display: inline-block; background: #2d1b1b;
        color: #f38ba8; border-radius: 4px;
        padding: 2px 8px; font-size: 0.75rem; margin: 2px;
    }

    /* Score colors */
    .score-high { color: #a6e3a1; font-weight: 700; }
    .score-mid  { color: #f9e2af; font-weight: 700; }
    .score-low  { color: #f38ba8; font-weight: 700; }

    /* Mobile : sidebar masquée par défaut, tout en colonne */
    @media (max-width: 768px) {
        .metric-value { font-size: 1.2rem; }
        .metric-label { font-size: 0.7rem; }
        .metric-card  { padding: 0.6rem; }
        .log-box      { height: 180px; font-size: 0.7rem; }
        .issue-chip   { font-size: 0.7rem; }

        /* Colonnes Streamlit en pleine largeur sur mobile */
        [data-testid="column"] {
            width: 100% !important;
            flex: 1 1 100% !important;
            min-width: 100% !important;
        }

        /* Padding réduit */
        .block-container {
            padding: 1rem 0.5rem !important;
        }

        /* Texte plus lisible */
        p, li, label { font-size: 0.9rem !important; }

        /* Bouton plus grand sur mobile */
        .stButton > button {
            padding: 0.8rem 1rem;
            font-size: 1.1rem;
        }
    }
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Session state init
# ---------------------------------------------------------------------------
def _init_state():
    defaults = {
        "running": False,
        "logs": [],
        "prospects": [],
        "log_queue": queue.Queue(),
        "run_done": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()

# Démarrage du thread d'envoi différé (idempotent — ne démarre qu'une fois par process)
from services import scheduler as _scheduler
_scheduler.ensure_running()

# Démarrage du suivi de réponses IMAP (démarré plus tard après lecture des credentials)

# Chargement des paramètres sauvegardés (fallback sur env vars, puis "")
from settings_manager import load_settings as _load_settings, save_settings as _save_settings
_saved = _load_settings()

def _get(key: str, env_var: str = "", default: str = "") -> str:
    """Priorité : settings.json > variable d'env > valeur par défaut."""
    return _saved.get(key) or os.getenv(env_var, "") or default


# ---------------------------------------------------------------------------
# Sidebar — Configuration
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown("## 🎯 Prospection B2B")
    st.markdown("---")

    st.markdown("### 📡 Source de prospection")
    from services.sources import SOURCE_LABELS as _SRC_LABELS
    source_type = st.selectbox(
        "Source",
        options=list(_SRC_LABELS.keys()),
        format_func=lambda x: _SRC_LABELS[x],
        index=list(_SRC_LABELS.keys()).index(_get("source_type") or "google_maps"),
        label_visibility="collapsed",
    )

    # Config spécifique à la source
    ft_client_id = ft_client_secret = google_cx = ""

    if source_type == "france_travail":
        st.markdown("**Client ID France Travail**")
        ft_client_id = st.text_input(
            "FT Client ID", value=_get("ft_client_id", "FT_CLIENT_ID"),
            placeholder="PAR_xxxx…", label_visibility="collapsed",
        )
        st.markdown("**Client Secret France Travail**")
        ft_client_secret = st.text_input(
            "FT Client Secret", type="password",
            value=_get("ft_client_secret", "FT_CLIENT_SECRET"),
            placeholder="xxxxxxxx-xxxx-…", label_visibility="collapsed",
        )
        with st.expander("ℹ️ Comment obtenir les identifiants France Travail ?"):
            st.markdown("""
1. Inscris-toi sur [francetravail.io](https://francetravail.io/inscription)
2. **Mes APIs** → **Référencer une nouvelle application**
3. Coche **Offres d'emploi v2** dans les API souhaitées
4. Récupère **Client ID** et **Client Secret** dans l'onglet **Mes applications**
""")

    elif source_type == "google_search":
        st.markdown("**Custom Search Engine ID (cx)**")
        google_cx = st.text_input(
            "CX", value=_get("google_cx", "GOOGLE_CX"),
            placeholder="017576…:xxxxxxx", label_visibility="collapsed",
        )
        with st.expander("ℹ️ Comment créer un moteur de recherche Google ?"):
            st.markdown("""
1. Va sur [programmablesearchengine.google.com](https://programmablesearchengine.google.com/controlpanel/all)
2. **Ajouter** → donne un nom → **Rechercher dans tout le web** ✓ → **Créer**
3. Copie l'**ID du moteur de recherche** (`017576…:xxx`) et colle-le ci-dessus
4. La clé Google API existante est réutilisée automatiquement
""")

    elif source_type == "sirene":
        st.caption("✅ Aucune clé requise — API officielle data.gouv.fr (gratuite).")

    elif source_type == "pages_jaunes":
        st.caption("✅ Aucune clé requise — scraping de l'annuaire public Pages Jaunes.")

    elif source_type == "linkedin_csv":
        st.caption("📎 Le fichier CSV LinkedIn s'importe dans la zone principale.")

    st.markdown("---")
    st.markdown("### 🔑 Clés API")

    google_key_required = source_type in ("google_maps", "google_search")
    st.markdown(
        "**Google Places / Search API Key** "
        "— [Obtenir ici ↗](https://console.cloud.google.com/apis/credentials)"
        + ("" if google_key_required else " *(optionnel pour cette source)*")
    )
    google_key = st.text_input(
        "Google Places API Key",
        type="password",
        value=_get("google_api_key", "GOOGLE_PLACES_API_KEY"),
        placeholder="AIzaSy...",
        label_visibility="collapsed",
    )
    with st.expander("ℹ️ Comment créer cette clé ?"):
        st.markdown("""
1. Va sur [Google Cloud Console](https://console.cloud.google.com/apis/credentials)
2. **Créer des identifiants** → **Clé API**
3. Dans **Bibliothèque d'API**, active :
   - *Places API* (obligatoire pour Google Maps)
   - *Custom Search JSON API* (pour la source Google Search)
   - *PageSpeed Insights API* (pour les scores de perf mobile)
4. Optionnel : restreins la clé à ces API (onglet **Restrictions de clé API**)
5. Copie la clé (`AIzaSy…`) et colle-la ci-dessus
""")

    st.markdown("**CRM**")
    crm_type = st.selectbox(
        "CRM", options=["Aucun", "Notion", "HubSpot"],
        label_visibility="collapsed",
    ).lower()
    if crm_type == "notion":
        crm_key = st.text_input(
            "Notion API Key", type="password",
            value=_get("notion_api_key", "NOTION_API_KEY"), placeholder="secret_...",
            label_visibility="collapsed",
        )
        crm_extra = {
            "database_id": st.text_input(
                "Notion Database ID", value=_get("notion_database_id", "NOTION_DATABASE_ID"),
                placeholder="c2507703-...", label_visibility="collapsed",
            )
        }
        with st.expander("ℹ️ Notion — comment créer la clé et le Database ID ?"):
            st.markdown("""
**Token d'intégration :**
1. Va sur [notion.so/my-integrations](https://www.notion.so/my-integrations)
2. **+ Nouvelle intégration** → nom "ProspectionBot" → Submit
3. Copie le **Token d'intégration interne** (`secret_…`)

**Database ID :**
1. Ouvre ta base Notion dans le navigateur
2. URL : `notion.so/MonEspace/`**`c2507703175647aaf2132a76c00e06`**`?v=…`
3. Le Database ID est la partie surlignée (32 car. après le dernier `/` avant `?v=`)
4. ⚠️ Invite l'intégration dans ta base : ouvre la base → **⋯** → **Connexions** → ajoute "ProspectionBot"
""")
    elif crm_type == "hubspot":
        crm_key = st.text_input(
            "HubSpot Private App Token", type="password",
            value=_get("hubspot_api_key", "HUBSPOT_API_KEY"), placeholder="pat-eu1-...",
            label_visibility="collapsed",
        )
        with st.expander("ℹ️ HubSpot — comment créer un token ?"):
            st.markdown("""
1. Dans HubSpot, va dans **Paramètres** (⚙️) → **Intégrations** → **Applications privées**
2. **Créer une application privée** → donne-lui un nom (ex: ProspectionBot)
3. Onglet **Portées** : coche `crm.objects.contacts.write` et `crm.objects.contacts.read`
4. **Créer l'application** → copie le token (`pat-eu1-…`)
""")
        crm_extra = {}
    else:
        crm_key  = ""
        crm_extra = {}

    # Garder notion_key pour compat avec les sections historique/relances
    notion_key = crm_key if crm_type == "notion" else ""

    st.markdown(
        "**Brevo API Key** "
        "— [Obtenir ici ↗](https://app.brevo.com/settings/keys/api)"
    )
    brevo_key = st.text_input(
        "Brevo API Key",
        type="password",
        value=_get("brevo_api_key", "BREVO_API_KEY"),
        placeholder="xsmtpsib-...",
        label_visibility="collapsed",
    )
    with st.expander("ℹ️ Comment créer cette clé ?"):
        st.markdown("""
1. Connecte-toi sur [app.brevo.com](https://app.brevo.com)
2. Clique sur ton **avatar** en haut à droite → **SMTP & API**
3. Onglet **API Keys** → **Générer une nouvelle clé API**
4. Donne-lui un nom → copie la clé (`xsmtpsib-…`)
""")

    st.markdown("---")
    st.markdown("### 📧 Gmail (optionnel)")

    st.markdown("**Adresse Gmail**")
    gmail_address = st.text_input(
        "Adresse Gmail",
        value=_get("gmail_address", "GMAIL_ADDRESS"),
        placeholder="toi@gmail.com",
        label_visibility="collapsed",
    )

    st.markdown(
        "**Mot de passe d'application** "
        "— [Générer ici ↗](https://myaccount.google.com/apppasswords) *(pas ton vrai mdp)*"
    )
    gmail_password = st.text_input(
        "Mot de passe d'application",
        type="password",
        value=os.getenv("GMAIL_APP_PASSWORD", ""),
        placeholder="xxxx xxxx xxxx xxxx",
        label_visibility="collapsed",
    )
    with st.expander("ℹ️ Comment créer un mot de passe d'application Google ?"):
        st.markdown("""
1. Va sur [myaccount.google.com/security](https://myaccount.google.com/security)
2. Active la **Validation en 2 étapes** si ce n'est pas déjà fait
3. Recherche **"Mots de passe des applications"** dans la barre de recherche de ton compte
4. Sélectionne **Autre (nom personnalisé)** → entre "ProspectionBot" → **Générer**
5. Copie le mot de passe à 16 caractères affiché (`xxxx xxxx xxxx xxxx`)
6. Entre TON adresse Gmail dans le champ "Adresse Gmail" ci-dessus
""")

    st.markdown("---")
    st.markdown("### 👤 Ta signature")
    your_name = st.text_input("Prénom", value=_get("your_name", "YOUR_NAME"))
    your_title = st.text_input("Titre", value=_get("your_title", "YOUR_TITLE"))
    your_email = st.text_input("Ton email", value=_get("your_email", "YOUR_EMAIL"))
    your_website = st.text_input("Ton site", value=_get("your_website", "YOUR_WEBSITE"))

    st.markdown("---")

    # Démarrage auto du suivi de réponses si credentials présents
    _rt_addr = _get("gmail_address", "GMAIL_ADDRESS")
    _rt_pwd  = os.getenv("GMAIL_APP_PASSWORD", "")
    if _rt_addr and _rt_pwd:
        from services import reply_tracker as _rt
        _rt.ensure_running(_rt_addr, _rt_pwd)

    st.caption("🔒 Tes clés restent sur ta machine. Rien n'est envoyé à l'extérieur.")


# ---------------------------------------------------------------------------
# Titre principal
# ---------------------------------------------------------------------------
from profiles import PROFILES, get_profile
from service_profiles import (
    SERVICE_PROFILES, SERVICE_CATEGORY_LABELS,
    get_service, list_services,
)
from target_segments import (
    TARGET_SEGMENTS, TARGET_SECTOR_LABELS, SIZE_LABELS,
    get_target, list_targets,
)

st.markdown("# 🎯 Prospection B2B Automatisée")
st.markdown("Trouve des prospects locaux, analyse leur besoin et génère des cold emails/SMS en un clic.")
st.markdown("---")

# ---------------------------------------------------------------------------
# Sélection service × cible
# ---------------------------------------------------------------------------
st.markdown("### 🧩 Votre activité")

# Sélecteur catégorie de service (radio horizontal)
_svc_cats = list(SERVICE_CATEGORY_LABELS.keys())
_saved_svc_cat = _get("service_category", _svc_cats[0])
_svc_cat_idx = _svc_cats.index(_saved_svc_cat) if _saved_svc_cat in _svc_cats else 0
selected_svc_cat = st.radio(
    "Catégorie",
    options=_svc_cats,
    format_func=lambda c: SERVICE_CATEGORY_LABELS[c],
    index=_svc_cat_idx,
    horizontal=True,
    label_visibility="collapsed",
    key="svc_cat_radio",
)

# Sélecteur service (filtré par catégorie)
_svcs_in_cat = [s for s in SERVICE_PROFILES if s.category == selected_svc_cat]
_svc_ids = [s.id for s in _svcs_in_cat]
_svc_by_id = {s.id: s for s in SERVICE_PROFILES}
_saved_svc = _get("service_id", _svc_ids[0] if _svc_ids else "web_refonte")
_svc_idx = _svc_ids.index(_saved_svc) if _saved_svc in _svc_ids else 0
selected_service_id = st.selectbox(
    "Service",
    options=_svc_ids,
    format_func=lambda sid: f"{_svc_by_id[sid].emoji} {_svc_by_id[sid].name}",
    index=_svc_idx,
    label_visibility="collapsed",
    key="service_selectbox",
)
selected_service = _svc_by_id[selected_service_id]
st.caption(f"*{selected_service.description}*")

st.markdown("---")
st.markdown("### 🎯 Votre cible")

# Sélecteur secteur cible (radio horizontal)
_tgt_sectors = list(TARGET_SECTOR_LABELS.keys())
_saved_tgt_sector = _get("target_sector", _tgt_sectors[0])
_tgt_sector_idx = _tgt_sectors.index(_saved_tgt_sector) if _saved_tgt_sector in _tgt_sectors else 0
selected_tgt_sector = st.radio(
    "Secteur",
    options=_tgt_sectors,
    format_func=lambda s: TARGET_SECTOR_LABELS[s],
    index=_tgt_sector_idx,
    horizontal=True,
    label_visibility="collapsed",
    key="tgt_sector_radio",
)

# Sélecteur cible (filtré par secteur)
_tgts_in_sector = [t for t in TARGET_SEGMENTS if t.sector == selected_tgt_sector]
_tgt_ids = [t.id for t in _tgts_in_sector]
_tgt_by_id = {t.id: t for t in TARGET_SEGMENTS}
_saved_tgt = _get("target_id", _tgt_ids[0] if _tgt_ids else "restaurants")
_tgt_idx = _tgt_ids.index(_saved_tgt) if _saved_tgt in _tgt_ids else 0
selected_target_id = st.selectbox(
    "Cible",
    options=_tgt_ids,
    format_func=lambda tid: f"{_tgt_by_id[tid].emoji} {_tgt_by_id[tid].name}  ·  {SIZE_LABELS[_tgt_by_id[tid].target_size]}",
    index=_tgt_idx,
    label_visibility="collapsed",
    key="target_selectbox",
)
selected_target = _tgt_by_id[selected_target_id]
st.caption(f"*{selected_target.description}*")

st.markdown("---")

# ---------------------------------------------------------------------------
# Critères de recherche — pré-remplis depuis le profil
# ---------------------------------------------------------------------------
st.markdown("### 📍 Configurez votre campagne")

col1, col2 = st.columns([2, 1])

with col1:
    location = st.text_input(
        "📌 Ville / Zone géographique",
        value=selected_target.location_default or os.getenv("SEARCH_LOCATION", "Lyon, France"),
        placeholder="Paris, France",
    )
    keywords_raw = st.text_area(
        "🔑 Mots-clés cibles (un par ligne)",
        value="\n".join(selected_target.keywords),
        height=150,
        placeholder="restaurant\nboulangerie\ncoiffeur",
    )
    keywords = [k.strip() for k in keywords_raw.splitlines() if k.strip()]

    from services.mailer import EmailStyle, EMAIL_STYLE_LABELS, build_dynamic_email
    from services.google_maps import Prospect as _PreviewProspect

    with st.expander("✉️ Style des emails", expanded=False):
        _email_intonation = st.radio(
            "Intonation",
            options=list(EMAIL_STYLE_LABELS["intonation"].keys()),
            format_func=lambda k: EMAIL_STYLE_LABELS["intonation"][k],
            index=list(EMAIL_STYLE_LABELS["intonation"].keys()).index(
                _get("email_intonation") if _get("email_intonation") in EMAIL_STYLE_LABELS["intonation"] else "professional"
            ),
            horizontal=True,
            key="email_intonation",
        )
        _email_length = st.radio(
            "Longueur",
            options=list(EMAIL_STYLE_LABELS["length"].keys()),
            format_func=lambda k: EMAIL_STYLE_LABELS["length"][k],
            index=list(EMAIL_STYLE_LABELS["length"].keys()).index(
                _get("email_length") if _get("email_length") in EMAIL_STYLE_LABELS["length"] else "medium"
            ),
            horizontal=True,
            key="email_length",
        )
        _email_salutation = st.selectbox(
            "Formule d'ouverture",
            options=list(EMAIL_STYLE_LABELS["salutation"].keys()),
            format_func=lambda k: EMAIL_STYLE_LABELS["salutation"][k],
            index=list(EMAIL_STYLE_LABELS["salutation"].keys()).index(
                _get("email_salutation") if _get("email_salutation") in EMAIL_STYLE_LABELS["salutation"] else "neutral"
            ),
            key="email_salutation",
        )
        _email_cta = st.selectbox(
            "Appel à l'action",
            options=list(EMAIL_STYLE_LABELS["cta"].keys()),
            format_func=lambda k: EMAIL_STYLE_LABELS["cta"][k],
            index=list(EMAIL_STYLE_LABELS["cta"].keys()).index(
                _get("email_cta") if _get("email_cta") in EMAIL_STYLE_LABELS["cta"] else "audit"
            ),
            key="email_cta",
        )

        # Prévisualisation avec un faux prospect
        _preview_style = EmailStyle(
            intonation=_email_intonation,
            length=_email_length,
            salutation=_email_salutation,
            cta=_email_cta,
        )
        st.markdown("**Aperçu (site sans HTTPS + pas de formulaire) :**")
        _preview_prospect = _PreviewProspect(
            place_id="preview",
            name="Votre Prospect",
            address="",
            phone=None,
            website="http://exemple.com",
            rating=None,
            user_ratings_total=0,
            keyword="",
            maps_url="",
        )
        _preview_prospect.issue_keys = ["https", "lead_form"]
        _preview_prospect.score = 40
        _preview_text = build_dynamic_email(
            _preview_prospect,
            _preview_style,
            your_name=_get("your_name", "YOUR_NAME") or "Votre Nom",
            your_title=_get("your_title", "YOUR_TITLE") or "",
            your_offer=selected_service.your_offer or "vous aider à améliorer votre présence en ligne",
        )
        st.code(_preview_text, language=None)

    st.markdown("**📱 Accroche SMS** *(max 160 caractères)*")
    sms_hook = st.text_input(
        "Accroche SMS",
        value=selected_service.sms_hook,
        label_visibility="collapsed",
    )
    if len(sms_hook) > 160:
        st.warning(f"⚠️ SMS trop long : {len(sms_hook)}/160 caractères")

    # Variables de compatibilité (toujours référencées ailleurs dans app.py)
    email_hook = selected_service.email_hook

with col2:
    st.markdown("**⚙️ Paramètres**")
    your_offer = st.text_area(
        "🎁 Mon offre",
        value=selected_service.your_offer,
        height=80,
        help="Décrivez votre offre en 1-2 phrases",
    )
    max_results = st.slider("Prospects par mot-clé", 1, 20, 5)
    min_rating = st.slider(
        "Note Google minimum ⭐",
        min_value=1.0, max_value=5.0, value=3.0, step=0.5,
        help="Les établissements en dessous de cette note sont ignorés (probablement en difficulté)",
    )
    _score_dir = selected_service.score_direction
    _score_default = (selected_target.score_threshold_override or selected_service.score_threshold_default)
    if _score_dir == "desc":
        score_threshold = st.slider(
            "Score min requis",
            min_value=0, max_value=100, value=_score_default,
            help="Score élevé = bonne opportunité. Ex coursier : 70 = exclure les établissements déjà bien couverts en livraison.",
        )
    else:
        score_threshold = st.slider(
            "Score max à contacter",
            min_value=0, max_value=100, value=_score_default,
            help="100 = tous les prospects. Baisse pour ne garder que les sites avec beaucoup de problèmes.",
        )
    radius = st.select_slider(
        "Rayon de recherche",
        options=[1000, 2000, 5000, 10000, 20000, 50000],
        value=10000,
        format_func=lambda x: f"{x//1000} km",
    )
    send_emails = st.toggle("📧 Envoyer les emails auto", value=False)
    if send_emails:
        st.warning("⚠️ Seuls les prospects avec un email trouvé recevront un mail.")
        _email_mode = st.radio(
            "Mode d'envoi", ["📤 Immédiat", "⏰ Programmé"],
            horizontal=True, label_visibility="collapsed",
        )
        if _email_mode == "⏰ Programmé":
            from datetime import date as _dt_date, timedelta as _dt_td, time as _dt_time
            _sched_date = st.date_input("Date d'envoi", value=_dt_date.today() + _dt_td(days=1), min_value=_dt_date.today())
            _sched_time = st.time_input("Heure d'envoi", value=_dt_time(9, 0))
        else:
            _sched_date = None
            _sched_time = None
    else:
        _email_mode = "📤 Immédiat"
        _sched_date = None
        _sched_time = None
    send_sms_toggle = st.toggle("📱 Envoyer les SMS auto", value=False)
    if send_sms_toggle:
        st.warning("⚠️ Seuls les numéros mobiles (06/07) recevront un SMS.")
    cache_ttl_days = st.slider(
        "⚡ Cache analyses (jours)", 1, 90, 30,
        help="Durée de validité : un site analysé il y a moins de X jours ne sera pas réanalysé.",
    )

# LinkedIn CSV uploader (visible seulement si source linkedin_csv sélectionnée)
linkedin_content = ""
if source_type == "linkedin_csv":
    st.markdown("---")
    st.markdown("### 📎 Import CSV LinkedIn")
    st.caption(
        "Exporte tes contacts depuis **LinkedIn Sales Navigator** (Accounts/Leads → Export) "
        "ou **Mes connexions** (Paramètres → Confidentialité → Obtenir une copie de tes données)."
    )
    _uploaded_csv = st.file_uploader(
        "Déposer le fichier CSV LinkedIn",
        type=["csv"],
        help="Format auto-détecté : Sales Navigator, Connexions, ou tout CSV avec colonnes Company/Email/Website.",
    )
    if _uploaded_csv:
        try:
            linkedin_content = _uploaded_csv.read().decode("utf-8-sig")
            _n_rows = max(0, len(linkedin_content.splitlines()) - 1)
            st.success(f"✅ {_n_rows} ligne(s) importée(s) — prêt à analyser.")
        except Exception as _e:
            st.error(f"❌ Erreur de lecture : {_e}")

st.markdown("---")


# ---------------------------------------------------------------------------
# Logger qui alimente la queue Streamlit
# ---------------------------------------------------------------------------
class QueueLogger:
    def __init__(self, q: queue.Queue):
        self.q = q

    def _emit(self, level: str, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.q.put(f"[{ts}] {level} {msg}")

    def info(self, msg: str, *a):    self._emit("ℹ️ ", msg % a if a else msg)
    def debug(self, msg: str, *a):   self._emit("🔍", msg % a if a else msg)
    def warning(self, msg: str, *a): self._emit("⚠️ ", msg % a if a else msg)
    def error(self, msg: str, *a):   self._emit("❌", msg % a if a else msg)
    def critical(self, msg: str, *a):self._emit("🔴", msg % a if a else msg)


# ---------------------------------------------------------------------------
# Thread de prospection
# ---------------------------------------------------------------------------
def run_prospection(params: dict, log_q: queue.Queue, result_container: list):
    """Tourne dans un thread séparé pour ne pas bloquer l'UI."""
    try:
        # Force les variables d'env AVANT tout import de config
        os.environ["GOOGLE_PLACES_API_KEY"] = params["google_key"]
        os.environ["NOTION_API_KEY"] = params["notion_key"]
        os.environ["BREVO_API_KEY"] = params["brevo_key"]
        os.environ["SEARCH_LOCATION"] = params["location"]
        os.environ["SEARCH_KEYWORDS"] = ",".join(params["keywords"])
        os.environ["SEARCH_RADIUS"] = str(params["radius"])
        os.environ["MAX_RESULTS_PER_KEYWORD"] = str(params["max_results"])
        os.environ["YOUR_NAME"] = params["your_name"]
        os.environ["YOUR_TITLE"] = params["your_title"]
        os.environ["YOUR_EMAIL"] = params["your_email"]
        os.environ["YOUR_WEBSITE"] = params["your_website"]
        os.environ["YOUR_OFFER"] = params.get("your_offer", "")
        os.environ["EMAIL_HOOK"] = params.get("email_hook", "")
        os.environ["SMS_HOOK"] = params.get("sms_hook", "")

        # Recharge la config avec les nouvelles valeurs
        import config as cfg_module
        from config import Config
        cfg_module.config = Config()
        c = cfg_module.config

        # Remplace le logger global par notre QueueLogger
        ui_logger = QueueLogger(log_q)
        cfg_module.logger = ui_logger

        import services.google_maps as gm_mod
        import services.analyzer as an_mod
        import services.mailer as ma_mod
        import services.notion_sync as no_mod
        gm_mod.logger = ui_logger
        an_mod.logger = ui_logger
        ma_mod.logger = ui_logger
        no_mod.logger = ui_logger

        # Recharge aussi le config dans chaque module
        gm_mod.config = c
        an_mod.config = c
        no_mod.config = c

        from services.google_maps import fetch_raw_candidates, build_prospect
        from services.analyzer import analyze_prospect
        from services.mailer import enrich_with_email, draft_email, EmailStyle as _EmailStyle
        from services.crm import get_exporter
        from history_manager import load_contacted_ids, mark_as_contacted
        from services import cache as _cache_mod
        from services.sources import (
            search_sirene, search_pages_jaunes,
            search_france_travail, search_google_custom, SOURCE_LABELS,
        )
        from services.sources.linkedin_csv import parse_linkedin_csv
        _cache_mod.set_ttl(params.get("cache_ttl_days", 30))

        os.makedirs(c.output_dir, exist_ok=True)

        target_per_kw   = params["max_results"]
        min_rating      = params.get("min_rating", 3.0)
        threshold       = params.get("contact_score_threshold", 100)
        score_direction = params.get("score_direction", "asc")
        weight_overrides = params.get("weight_overrides", {})
        already_contacted = load_contacted_ids()
        source           = params.get("source_type", "google_maps")

        all_qualified: list = []
        seen: set = set()
        workers = params.get("analysis_workers", 5)

        def _analyse_and_filter(candidates: list, label: str) -> list:
            """Analyse un lot de prospects et filtre par score. Retourne la liste qualifiée."""
            if not candidates:
                return []
            batch = candidates[:target_per_kw * 3]
            with ThreadPoolExecutor(max_workers=min(workers, len(batch))) as ex:
                analyzed = list(ex.map(partial(analyze_prospect, weight_overrides=weight_overrides), batch))
            qualified = []
            for p in analyzed:
                qualifies = (p.score >= threshold if score_direction == "desc" else p.score <= threshold)
                if qualifies:
                    qualified.append(p)
                    log_q.put(f"[--] ✅ {p.name} — score {p.score}/100")
            return qualified

        def _dedup(candidates: list) -> list:
            """Retire les prospects déjà contactés ou vus dans cette session."""
            out = []
            for p in candidates:
                if p.place_id in seen or p.place_id in already_contacted:
                    continue
                seen.add(p.place_id)
                out.append(p)
            return out

        # ── LinkedIn CSV : traitement hors boucle mots-clés ──────────────────
        if source == "linkedin_csv":
            csv_content = params.get("linkedin_content", "")
            if not csv_content:
                log_q.put("[--] ❌ Aucun fichier CSV LinkedIn fourni.")
            else:
                raw_li = parse_linkedin_csv(csv_content, "LinkedIn")
                li_ok = _dedup(raw_li)
                log_q.put(f"[--] 📎 {len(li_ok)} contact(s) LinkedIn à analyser…")
                all_qualified.extend(_analyse_and_filter(li_ok, "LinkedIn"))

        else:
            # ── Boucle par mot-clé (Google Maps + Sirene + PJ + FT + GS) ────
            for kw in params["keywords"]:
                src_label = SOURCE_LABELS.get(source, source)
                log_q.put(f"[--] 🔍 [{src_label}] '{kw}' — objectif {target_per_kw}…")

                if source == "google_maps":
                    # ── Phase 1 : Text Search ──
                    raw_candidates = fetch_raw_candidates(kw)
                    if not raw_candidates:
                        log_q.put(f"[--] ❌ Aucun résultat Google pour '{kw}'.")
                        continue

                    skip_contacted = skip_seen = skip_api = skip_rating = 0
                    raw_to_build: list = []
                    for raw in raw_candidates:
                        if len(raw_to_build) >= target_per_kw * 4:
                            break
                        pid = raw.get("place_id", "")
                        if not pid:
                            continue
                        if pid in seen:
                            skip_seen += 1; continue
                        if pid in already_contacted:
                            skip_contacted += 1; continue
                        seen.add(pid)
                        raw_to_build.append(raw)

                    if not raw_to_build:
                        log_q.put(f"[--] ⚠️  0/{target_per_kw} — tous déjà contactés ou vus.")
                        continue

                    # ── Phase 2 : Place Details en parallèle ──
                    with ThreadPoolExecutor(max_workers=min(workers, len(raw_to_build))) as ex:
                        built_list = list(ex.map(partial(build_prospect, keyword=kw), raw_to_build))

                    candidates: list = []
                    for p in built_list:
                        if p is None:
                            skip_api += 1
                        elif p.rating is not None and p.rating < min_rating:
                            skip_rating += 1
                        else:
                            candidates.append(p)

                    if not candidates:
                        reasons = []
                        if skip_api:    reasons.append(f"{skip_api} erreur(s) API")
                        if skip_rating: reasons.append(f"{skip_rating} note(s) trop basse(s)")
                        log_q.put(f"[--] ⚠️  0/{target_per_kw} — {' | '.join(reasons) or 'Google épuisé'}.")
                        continue

                else:
                    # ── Sources alternatives : retournent directement des Prospects ──
                    if source == "sirene":
                        raw = search_sirene(kw, params["location"], target_per_kw * 3)
                    elif source == "pages_jaunes":
                        raw = search_pages_jaunes(kw, params["location"], target_per_kw * 3)
                    elif source == "france_travail":
                        raw = search_france_travail(
                            kw, params["location"], target_per_kw * 3,
                            client_id=params.get("ft_client_id", ""),
                            client_secret=params.get("ft_client_secret", ""),
                        )
                    elif source == "google_search":
                        raw = search_google_custom(
                            kw, params["location"], target_per_kw * 3,
                            cx=params.get("google_cx", ""),
                        )
                    else:
                        raw = []

                    candidates = _dedup(raw)
                    if not candidates:
                        log_q.put(f"[--] ⚠️  0 résultat {src_label} pour '{kw}'.")
                        continue

                # ── Phase 3+4 : Analyse + filtre score (commun toutes sources) ──
                kw_qualified = _analyse_and_filter(candidates, kw)[:target_per_kw]
                log_q.put(f"[--] {'✅' if len(kw_qualified) >= target_per_kw else '⚠️ '} {len(kw_qualified)}/{target_per_kw} qualifiés pour '{kw}'.")
                all_qualified.extend(kw_qualified)

        all_prospects = all_qualified
        log_q.put(f"[--] 📋 {len(all_prospects)} prospect(s) qualifiés au total.")

        # Emails
        style_dict = params.get("email_style", {})
        _email_style = _EmailStyle(
            intonation=style_dict.get("intonation", "professional"),
            length=style_dict.get("length", "medium"),
            salutation=style_dict.get("salutation", "neutral"),
            cta=style_dict.get("cta", "audit"),
        )
        for _p in all_prospects:
            _p.email_draft = draft_email(_p, style=_email_style)
        all_prospects = list(all_prospects)

        # 4. Tri
        reverse_sort = (score_direction == "desc")
        all_prospects.sort(key=lambda p: p.score, reverse=reverse_sort)

        # 5. Sauvegarde locale
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = os.path.join(c.output_dir, f"prospects_{ts}.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump([p.to_dict() for p in all_prospects], f, ensure_ascii=False, indent=2)

        # 6. Export CRM
        crm_exporter = get_exporter(
            params.get("crm_type", "aucun"),
            params.get("crm_key", ""),
            **params.get("crm_extra", {}),
        )
        notion_page_ids: dict = {}
        if crm_exporter:
            crm_exporter.export(all_prospects)
            if hasattr(crm_exporter, "_last_exported_ids"):
                notion_page_ids = crm_exporter._last_exported_ids

        # 7. Gmail
        if params["send_emails"] and params["gmail_address"] and params["gmail_password"]:
            if params.get("email_send_mode") == "⏰ Programmé" and params.get("sched_date"):
                from datetime import datetime as _dtime
                from services import scheduler as _sched_mod
                _send_at = _dtime.strptime(
                    f"{params['sched_date']} {params.get('sched_time', '09:00')}",
                    "%Y-%m-%d %H:%M",
                ).timestamp()
                _n_sched = 0
                for _p in all_prospects:
                    if not _p.email or not _p.email_draft:
                        continue
                    _sched_mod.add_pending(
                        place_id=_p.place_id,
                        name=_p.name,
                        email=_p.email,
                        draft=_p.email_draft,
                        gmail_address=params["gmail_address"],
                        gmail_password=params["gmail_password"],
                        send_at=_send_at,
                        notion_page_id=notion_page_ids.get(_p.place_id, ""),
                        notion_api_key=params.get("crm_key", "") if params.get("crm_type") == "notion" else "",
                    )
                    _n_sched += 1
                log_q.put(
                    f"[--] ⏰ {_n_sched} email(s) programmé(s) pour le "
                    f"{params['sched_date']} à {params.get('sched_time', '09:00')}."
                )
            else:
                from services.gmail import send_all
                send_all(all_prospects, params["gmail_address"], params["gmail_password"])
                if notion_page_ids and params.get("crm_type") == "notion" and params.get("crm_key"):
                    from services.crm.notion import NotionExporter
                    _nu = NotionExporter(params["crm_key"], params.get("crm_extra", {}).get("database_id", ""))
                    for _p in all_prospects:
                        if _p.email:
                            _pid = notion_page_ids.get(_p.place_id)
                            if _pid:
                                _nu.update_status(_pid, "contacté")

        # 8. SMS Brevo
        if params["send_sms"] and params["brevo_key"]:
            from services.sms import send_all_sms
            send_all_sms(all_prospects)

        # 9. Marquage des prospects contactés (avec infos complètes pour les relances)
        mark_as_contacted(all_prospects, notion_page_ids=notion_page_ids)

        # 10. Historique
        from history_manager import save_run
        emails_found = sum(1 for p in all_prospects if p.email)
        mobiles_found = sum(1 for p in all_prospects if p.phone and (
            p.phone.replace(" ", "").startswith("06") or
            p.phone.replace(" ", "").startswith("07")
        ))
        save_run(
            profile_name=params.get("profile_name", "Custom"),
            location=params["location"],
            keywords=params["keywords"],
            total=len(all_prospects),
            no_site=sum(1 for p in all_prospects if not p.has_website()),
            emails_found=emails_found,
            mobiles_found=mobiles_found,
            output_file=json_path,
        )

        result_container.extend(all_prospects)
        log_q.put("__DONE__")

    except Exception as exc:
        log_q.put(f"[--] ❌ Erreur critique : {exc}")
        log_q.put("__DONE__")


# ---------------------------------------------------------------------------
# Bouton de lancement
# ---------------------------------------------------------------------------
col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
with col_btn2:
    _launch_disabled = st.session_state.running
    if source_type == "google_maps":
        _launch_disabled = _launch_disabled or not google_key or not keywords or not location
    elif source_type == "google_search":
        _launch_disabled = _launch_disabled or not google_key or not google_cx or not keywords or not location
    elif source_type == "france_travail":
        _launch_disabled = _launch_disabled or not ft_client_id or not ft_client_secret or not keywords or not location
    elif source_type == "linkedin_csv":
        _launch_disabled = _launch_disabled or not linkedin_content
    else:  # sirene, pages_jaunes
        _launch_disabled = _launch_disabled or not keywords or not location

    launch = st.button("🚀 Lancer la prospection", disabled=_launch_disabled)

if source_type == "google_maps" and not google_key:
    st.info("👈 Renseigne ta clé Google Places dans la barre latérale pour commencer.")
elif source_type == "google_search" and not google_cx:
    st.info("👈 Renseigne ton Custom Search Engine ID (cx) dans la barre latérale.")
elif source_type == "france_travail" and (not ft_client_id or not ft_client_secret):
    st.info("👈 Renseigne tes identifiants France Travail dans la barre latérale.")
elif source_type == "linkedin_csv" and not linkedin_content:
    st.info("👆 Importe un fichier CSV LinkedIn ci-dessus pour commencer.")

# ---------------------------------------------------------------------------
# Démarrage du thread
# ---------------------------------------------------------------------------
if launch and not st.session_state.running:
    st.session_state.running = True
    st.session_state.run_done = False
    st.session_state.logs = []
    st.session_state.prospects = []
    st.session_state.log_queue = queue.Queue()

    # Persistance des paramètres (rechargés comme defaults au prochain démarrage)
    _save_settings({
        "google_api_key":    google_key,
        "source_type":       source_type,
        "ft_client_id":      ft_client_id or None,
        "ft_client_secret":  ft_client_secret or None,
        "google_cx":         google_cx or None,
        "crm_type":          crm_type,
        "notion_api_key":    crm_key if crm_type == "notion" else None,
        "notion_database_id": crm_extra.get("database_id") if crm_type == "notion" else None,
        "hubspot_api_key":   crm_key if crm_type == "hubspot" else None,
        "brevo_api_key":     brevo_key,
        "gmail_address":     gmail_address,
        "your_name":         your_name,
        "your_title":        your_title,
        "your_email":        your_email,
        "your_website":      your_website,
        "service_id":        selected_service_id,
        "service_category":  selected_svc_cat,
        "target_id":         selected_target_id,
        "target_sector":     selected_tgt_sector,
        "email_intonation":  _email_intonation,
        "email_length":      _email_length,
        "email_salutation":  _email_salutation,
        "email_cta":         _email_cta,
    })

    result_container = []

    params = {
        "google_key": google_key,
        "notion_key": notion_key,
        "crm_type":   crm_type,
        "crm_key":    crm_key,
        "crm_extra":  crm_extra,
        "brevo_key": brevo_key,
        "location": location,
        "keywords": keywords,
        "radius": radius,
        "max_results": max_results,
        "your_name": your_name,
        "your_title": your_title or selected_service.your_title,
        "your_email": your_email,
        "your_website": your_website,
        "your_offer": your_offer,
        "email_hook": email_hook,
        "sms_hook": sms_hook,
        "email_style": {
            "intonation": _email_intonation,
            "length":     _email_length,
            "salutation": _email_salutation,
            "cta":        _email_cta,
        },
        "profile_id": f"{selected_service_id}_x_{selected_target_id}",
        "profile_name": f"{selected_service.emoji} {selected_service.name}  →  {selected_target.emoji} {selected_target.name}",
        "weight_overrides": selected_service.check_weight_overrides,
        "score_direction": selected_service.score_direction,
        "min_rating": min_rating,
        "contact_score_threshold": score_threshold,
        "analysis_workers": int(os.getenv("ANALYSIS_WORKERS", "5")),
        "send_emails": send_emails,
        "gmail_address": gmail_address,
        "gmail_password": gmail_password,
        "send_sms": send_sms_toggle,
        "cache_ttl_days": cache_ttl_days,
        "email_send_mode": _email_mode,
        "sched_date": _sched_date.isoformat() if _sched_date else None,
        "sched_time": _sched_time.strftime("%H:%M") if _sched_time else None,
        "source_type":       source_type,
        "ft_client_id":      ft_client_id,
        "ft_client_secret":  ft_client_secret,
        "google_cx":         google_cx,
        "linkedin_content":  linkedin_content,
    }

    thread = threading.Thread(
        target=run_prospection,
        args=(params, st.session_state.log_queue, result_container),
        daemon=True,
    )
    thread.start()
    st.session_state._thread = thread
    st.session_state._results = result_container

# ---------------------------------------------------------------------------
# Affichage live des logs
# ---------------------------------------------------------------------------
if st.session_state.running or st.session_state.run_done:
    st.markdown("### 📡 Logs en temps réel")
    log_placeholder = st.empty()
    status_placeholder = st.empty()

    # Vide la queue dans la liste de logs
    q = st.session_state.log_queue
    while not q.empty():
        msg = q.get_nowait()
        if msg == "__DONE__":
            st.session_state.running = False
            st.session_state.run_done = True
            if hasattr(st.session_state, "_results"):
                st.session_state.prospects = list(st.session_state._results)
        else:
            st.session_state.logs.append(msg)

    # Affiche les logs
    log_html = "<div class='log-box'>" + "<br>".join(
        st.session_state.logs[-100:]
    ) + "</div>"
    log_placeholder.markdown(log_html, unsafe_allow_html=True)

    if st.session_state.running:
        status_placeholder.info("⏳ Prospection en cours…")
        time.sleep(1)
        st.rerun()
    else:
        status_placeholder.success("✅ Prospection terminée !")


# ---------------------------------------------------------------------------
# Résultats
# ---------------------------------------------------------------------------
if st.session_state.prospects:
    prospects = st.session_state.prospects
    st.markdown("---")
    st.markdown("## 📊 Résultats")

    # Métriques
    no_site     = sum(1 for p in prospects if not p.has_website())
    critical    = sum(1 for p in prospects if p.score < 40)
    avg_score   = int(sum(p.score for p in prospects) / len(prospects))
    emails_ok   = sum(1 for p in prospects if p.email)
    mobiles_ok  = sum(1 for p in prospects if p.phone and (
        p.phone.replace(" ", "").startswith("06") or
        p.phone.replace(" ", "").startswith("07")
    ))

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    for col, value, label, color in [
        (m1, len(prospects),  "Total prospects",    "#667eea"),
        (m2, no_site,         "Sans site 🔴",        "#f38ba8"),
        (m3, critical,        "Score < 40 🟡",       "#fab387"),
        (m4, avg_score,       "Score moyen",         "#667eea"),
        (m5, emails_ok,       "📧 Emails trouvés",   "#a6e3a1"),
        (m6, mobiles_ok,      "📱 Mobiles trouvés",  "#a6e3a1"),
    ]:
        with col:
            st.markdown(f"""<div class='metric-card'>
                <div class='metric-value' style='color:{color}'>{value}</div>
                <div class='metric-label'>{label}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("")

    # Filtre
    col_f1, col_f2 = st.columns([2, 1])
    with col_f1:
        filter_opt = st.radio(
            "Afficher :",
            ["Tous", "Sans site uniquement", "Email trouvé", "Mobile trouvé", "Score < 40"],
            horizontal=True,
        )
    with col_f2:
        sort_opt = st.selectbox("Trier par :", ["Opportunité (score ↑)", "Nom (A→Z)", "Note Google (↓)"])

    # Application des filtres
    filtered = prospects
    if filter_opt == "Sans site uniquement":
        filtered = [p for p in prospects if not p.has_website()]
    elif filter_opt == "Email trouvé":
        filtered = [p for p in prospects if p.email]
    elif filter_opt == "Mobile trouvé":
        filtered = [p for p in prospects if p.phone and (
            p.phone.replace(" ", "").startswith("06") or
            p.phone.replace(" ", "").startswith("07")
        )]
    elif filter_opt == "Score < 40":
        filtered = [p for p in prospects if p.score < 40]

    if sort_opt == "Nom (A→Z)":
        filtered = sorted(filtered, key=lambda p: p.name)
    elif sort_opt == "Note Google (↓)":
        filtered = sorted(filtered, key=lambda p: p.rating or 0, reverse=True)

    st.markdown(f"### 🏆 {len(filtered)} prospect(s) — triés par {sort_opt.lower()}")

    for p in filtered:
        score_emoji = "🟢" if p.score >= 70 else ("🟡" if p.score >= 40 else "🔴")
        email_badge = "📧✅" if p.email else "📧❌"
        phone_type = ""
        if p.phone:
            num = p.phone.replace(" ", "")
            phone_type = "📱" if (num.startswith("06") or num.startswith("07")) else "☎️"

        header = f"{score_emoji} **{p.name}** — Score {p.score}/100 — {email_badge} {phone_type}"
        with st.expander(header):
            c1, c2 = st.columns([1, 1])
            with c1:
                st.markdown(f"**📍 Adresse :** {p.address}")
                # Téléphone avec badge mobile/fixe
                if p.phone:
                    num = p.phone.replace(" ", "")
                    is_mobile = num.startswith("06") or num.startswith("07")
                    badge = "📱 Mobile" if is_mobile else "☎️ Fixe"
                    st.markdown(f"**Téléphone :** {p.phone} — `{badge}`")
                else:
                    st.markdown("**Téléphone :** —")

                # Email avec statut
                if p.email:
                    st.markdown(f"**📧 Email trouvé :** `{p.email}`")
                else:
                    st.markdown("**📧 Email :** non trouvé sur le site")

                # Site web + CMS détecté
                if p.website:
                    cms_badge = f" `{p.cms}`" if p.cms else ""
                    st.markdown(f"**🌐 Site :** [{p.website}]({p.website}){cms_badge}")
                else:
                    st.markdown("**🌐 Site :** ❌ Aucun site web")

                st.markdown(f"**🔑 Mot-clé :** `{p.keyword}`")
                if p.rating:
                    stars = "⭐" * round(p.rating)
                    st.markdown(f"**Note Google :** {stars} {p.rating}/5 ({p.user_ratings_total} avis)")
                if p.maps_url:
                    st.markdown(f"[📌 Voir sur Google Maps]({p.maps_url})")

            with c2:
                st.markdown("**🔬 Problèmes détectés :**")
                if p.issues:
                    for issue in p.issues:
                        short = issue.split("→")[0].strip()
                        st.markdown(f"<span class='issue-chip'>⚠️ {short}</span>", unsafe_allow_html=True)
                else:
                    st.markdown("✅ Aucun problème majeur détecté")

            st.markdown("**✉️ Brouillon cold email :**")
            st.code(p.email_draft, language=None)

    # ---------------------------------------------------------------------------
    # Export
    # ---------------------------------------------------------------------------
    st.markdown("---")
    st.markdown("### 💾 Export")

    import csv, io
    col_e1, col_e2, col_e3 = st.columns(3)

    with col_e1:
        json_data = json.dumps([p.to_dict() for p in filtered], ensure_ascii=False, indent=2)
        st.download_button(
            label="⬇️ Télécharger JSON",
            data=json_data,
            file_name=f"prospects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            use_container_width=True,
        )

    with col_e2:
        csv_buffer = io.StringIO()
        fieldnames = ["name", "keyword", "address", "phone", "email", "website",
                      "cms", "rating", "score", "issues_count", "issues_summary", "maps_url"]
        writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
        writer.writeheader()
        for p in filtered:
            writer.writerow({
                "name": p.name,
                "keyword": p.keyword,
                "address": p.address,
                "phone": p.phone or "",
                "email": p.email or "",
                "website": p.website or "",
                "cms": p.cms or "",
                "rating": p.rating or "",
                "score": p.score,
                "issues_count": len(p.issues),
                "issues_summary": " | ".join(p.issues[:3]),
                "maps_url": p.maps_url,
            })
        st.download_button(
            label="⬇️ Télécharger CSV",
            data=csv_buffer.getvalue().encode("utf-8-sig"),
            file_name=f"prospects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_e3:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Prospects"

            headers = ["Nom", "Mot-clé", "Adresse", "Téléphone", "Email",
                       "Site", "CMS", "Note ⭐", "Score", "Nb problèmes", "Problèmes (top 3)", "Google Maps"]
            col_widths = [30, 15, 40, 15, 32, 40, 12, 8, 8, 12, 70, 50]

            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.row_dimensions[1].height = 20

            fill_green  = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            fill_yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            fill_red    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

            for ri, p in enumerate(filtered, 2):
                score_fill = fill_green if p.score >= 70 else (fill_yellow if p.score >= 40 else fill_red)
                row_vals = [
                    p.name, p.keyword, p.address, p.phone or "",
                    p.email or "", p.website or "", p.cms or "",
                    p.rating or "", p.score, len(p.issues),
                    " | ".join(p.issues[:3]), p.maps_url,
                ]
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=(ci == 11))
                    if ci == 9:  # Score
                        cell.fill = score_fill
                        cell.font = Font(bold=True)

            ws.freeze_panes = "A2"

            _xls_buf = io.BytesIO()
            wb.save(_xls_buf)
            st.download_button(
                label="⬇️ Télécharger Excel",
                data=_xls_buf.getvalue(),
                file_name=f"prospects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except ImportError:
            st.button("⬇️ Excel — installe openpyxl", disabled=True, use_container_width=True)
            st.caption("`pip install openpyxl`")

# ---------------------------------------------------------------------------
# Sauvegarde profil custom
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("💾 Sauvegarder ce profil pour une prochaine fois"):
    save_name = st.text_input("Nom du profil", placeholder="Ex: Mon profil Lyon Dev Web")
    if st.button("💾 Sauvegarder") and save_name:
        from profile_manager import save_custom_profile
        from profiles import Profile
        import re, uuid
        custom_id = "custom_" + re.sub(r"[^a-z0-9]", "_", save_name.lower())[:20]
        new_profile = Profile(
            id=custom_id,
            emoji="⭐",
            name=save_name,
            description=f"Profil personnalisé — {location}",
            keywords=keywords,
            location=location,
            your_title=your_title,
            your_offer=your_offer,
            email_hook=email_hook,
            sms_hook=sms_hook,
            qualification_criteria=[],
            radius=radius,
            max_results=max_results,
        )
        save_custom_profile(new_profile)
        st.success(f"✅ Profil « {save_name} » sauvegardé ! Il apparaîtra dans la liste au prochain lancement.")

# ---------------------------------------------------------------------------
# Dashboard de statistiques
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("📊 Dashboard — Statistiques globales"):
    from history_manager import load_history as _lh, _load_contacted_data as _lcd2
    _hist = _lh()
    _cdata2 = _lcd2()

    if not _hist:
        st.info("Lancez au moins une campagne pour voir les statistiques.")
    else:
        _total_runs      = len(_hist)
        _total_prospects = sum(r.get("total_prospects", 0) for r in _hist)
        _total_emails    = sum(r.get("emails_trouvés", 0)  for r in _hist)
        _total_mobiles   = sum(r.get("mobiles_trouvés", 0) for r in _hist)
        _total_responded = sum(1 for v in _cdata2.values() if v.get("responded"))
        _total_contacted = len(_cdata2)

        # Métriques globales
        dc1, dc2, dc3, dc4, dc5 = st.columns(5)
        dc1.metric("Campagnes", _total_runs)
        dc2.metric("Prospects total", _total_prospects)
        dc3.metric("Emails trouvés", _total_emails)
        dc4.metric("Mobiles trouvés", _total_mobiles)
        _rrate = f"{_total_responded / _total_contacted * 100:.0f}%" if _total_contacted else "—"
        dc5.metric("Taux de réponse", _rrate)

        st.markdown("---")

        # Graphique : prospects + emails par campagne (10 dernières)
        try:
            import pandas as _pd

            _runs_data = [{
                "Campagne": r["date"][:10],
                "Prospects": r.get("total_prospects", 0),
                "Emails": r.get("emails_trouvés", 0),
                "Mobiles": r.get("mobiles_trouvés", 0),
            } for r in reversed(_hist[:10])]
            _df_runs = _pd.DataFrame(_runs_data).set_index("Campagne")
            st.markdown("**Prospects et emails par campagne (10 dernières)**")
            st.bar_chart(_df_runs[["Prospects", "Emails"]])

            # Top mots-clés
            _kw_counts: dict = {}
            for r in _hist:
                for kw in r.get("keywords", []):
                    _kw_counts[kw] = _kw_counts.get(kw, 0) + r.get("total_prospects", 0)
            if _kw_counts:
                _top_kw = sorted(_kw_counts.items(), key=lambda x: x[1], reverse=True)[:10]
                _df_kw = _pd.DataFrame(_top_kw, columns=["Mot-clé", "Prospects"]).set_index("Mot-clé")
                st.markdown("**Top mots-clés (par nombre de prospects cumulés)**")
                st.bar_chart(_df_kw)

            # Distribution des scores de la dernière campagne
            if st.session_state.prospects:
                _scores = [p.score for p in st.session_state.prospects]
                _bins = {"0-20": 0, "21-40": 0, "41-60": 0, "61-80": 0, "81-100": 0}
                for s in _scores:
                    if s <= 20:    _bins["0-20"]   += 1
                    elif s <= 40:  _bins["21-40"]  += 1
                    elif s <= 60:  _bins["41-60"]  += 1
                    elif s <= 80:  _bins["61-80"]  += 1
                    else:          _bins["81-100"] += 1
                _df_score = _pd.DataFrame(list(_bins.items()), columns=["Score", "Nombre"]).set_index("Score")
                st.markdown("**Distribution des scores (campagne en cours)**")
                st.bar_chart(_df_score)

        except ImportError:
            st.caption("pandas non disponible — install `pandas` pour les graphiques.")

# ---------------------------------------------------------------------------
# Test A/B — Statistiques de templates
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("🧪 Test A/B — Performance des templates email"):
    from history_manager import get_ab_stats
    ab = get_ab_stats()
    st.caption("Les prospects sont répartis 50/50 entre le template A (narratif) et le template B (court/direct).")
    col_a, col_b = st.columns(2)
    for col, variant, label in [(col_a, "A", "Template A — Narratif"), (col_b, "B", "Template B — Direct")]:
        s = ab.get(variant, {"total": 0, "responded": 0})
        rate = f"{s['responded']/s['total']*100:.0f}%" if s["total"] else "—"
        with col:
            st.markdown(f"**{label}**")
            st.metric("Envoyés", s["total"])
            st.metric("Réponses", s["responded"])
            st.metric("Taux de réponse", rate)

# ---------------------------------------------------------------------------
# Historique des campagnes
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("🕐 Historique des campagnes"):
    from history_manager import load_history
    history = load_history()
    if not history:
        st.info("Aucune campagne lancée pour l'instant.")
    else:
        for run in history:
            kw_str = ", ".join(run.get("keywords", [])[:3])
            st.markdown(
                f"**{run['date']}** — {run['profile']} — {run['location']} — "
                f"`{kw_str}` — "
                f"**{run['total_prospects']}** prospects | "
                f"📧 {run['emails_trouvés']} emails | "
                f"📱 {run['mobiles_trouvés']} mobiles"
            )
            st.divider()

# ---------------------------------------------------------------------------
# Historique
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("🗂️ Historique des contacts"):
    from history_manager import load_contacted_ids
    contacted = load_contacted_ids()
    st.write(f"**{len(contacted)}** établissement(s) déjà contacté(s) (ignorés aux prochains runs).")
    if contacted:
        if st.button("🗑️ Réinitialiser l'historique", type="secondary"):
            import json as _json
            history_path = os.path.join("output", "contacted_place_ids.json")
            if os.path.exists(history_path):
                os.remove(history_path)
            st.success("Historique effacé. Le prochain run reprospecttra depuis zéro.")
            st.rerun()

# ---------------------------------------------------------------------------
# Emails programmés
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("📬 Emails programmés"):
    from services import scheduler as _sched_ui
    _stats = _sched_ui.get_stats()
    col_s1, col_s2, col_s3 = st.columns(3)
    col_s1.metric("En attente", _stats["pending"])
    col_s2.metric("Envoyés", _stats["sent"])
    col_s3.metric("Total", _stats["total"])
    if _stats["pending"] > 0:
        st.info(f"⏰ {_stats['pending']} email(s) en attente d'envoi — le thread vérifie toutes les 60 secondes.")
    elif _stats["total"] == 0:
        st.caption("Aucun email programmé pour l'instant.")

# ---------------------------------------------------------------------------
# Suivi de réponses
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("📬 Suivi des réponses (IMAP)"):
    from services import reply_tracker as _rt_ui
    _rt_running = _rt_ui.is_running()
    if _rt_running:
        st.success("✅ Suivi actif — vérifie les réponses Gmail toutes les 5 minutes.")
    elif gmail_address and gmail_password:
        _rt_ui.ensure_running(gmail_address, gmail_password)
        st.info("⏳ Thread de suivi en cours de démarrage…")
    else:
        st.info("💡 Renseigne ton adresse Gmail et ton mot de passe d'application pour activer le suivi automatique des réponses.")
    from history_manager import _load_contacted_data as _lcd
    _cdata = _lcd()
    _responded = sum(1 for v in _cdata.values() if v.get("responded"))
    _total_c   = len(_cdata)
    if _total_c:
        col_rt1, col_rt2 = st.columns(2)
        col_rt1.metric("Prospects contactés", _total_c)
        col_rt2.metric("Réponses reçues", _responded)

# ---------------------------------------------------------------------------
# Cache d'analyse
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("⚡ Cache d'analyse (performances)"):
    from services import cache as _analysis_cache
    st.caption(
        "Les analyses récentes sont mises en cache pour éviter de refaire "
        "les appels HTTP et PageSpeed pour les mêmes sites."
    )
    n_cached = _analysis_cache.count()
    col_c1, col_c2 = st.columns([3, 1])
    with col_c1:
        st.write(f"**{n_cached}** site(s) actuellement en cache.")
    with col_c2:
        if st.button("🗑️ Vider", key="clear_cache", use_container_width=True, disabled=(n_cached == 0)):
            deleted = _analysis_cache.clear_all()
            st.success(f"✅ {deleted} entrée(s) supprimée(s).")
            st.rerun()

# ---------------------------------------------------------------------------
# Relances
# ---------------------------------------------------------------------------
st.markdown("---")
with st.expander("🔄 Relances — contacts sans réponse"):
    from history_manager import get_due_followups, mark_as_responded, mark_followup_sent
    followup_delay = int(os.getenv("FOLLOWUP_DELAY_DAYS", "5"))
    due = get_due_followups(followup_delay)

    if not due:
        st.success(f"✅ Aucun contact à relancer (seuil : {followup_delay} jours sans réponse).")
    else:
        st.info(f"**{len(due)} contact(s)** à relancer — contactés il y a plus de {followup_delay} jours sans réponse.")

        # Bouton pour générer tous les emails de relance d'un coup
        if st.button("📝 Générer tous les emails de relance", key="gen_followup"):
            from services.google_maps import Prospect as P
            from services.mailer import draft_followup_email
            drafts = []
            for contact in due:
                p = P(
                    place_id=contact["place_id"],
                    name=contact["name"],
                    address="",
                    phone=None,
                    website=None,
                    rating=None,
                    user_ratings_total=0,
                    keyword="",
                    email=contact.get("email") or None,
                )
                drafts.append((p.name, draft_followup_email(p), contact["place_id"]))
                mark_followup_sent(contact["place_id"])
                if crm_type == "notion" and crm_key:
                    from history_manager import get_notion_page_id
                    from services.crm.notion import NotionExporter
                    _np = get_notion_page_id(contact["place_id"])
                    if _np:
                        NotionExporter(crm_key, crm_extra.get("database_id", "")).update_status(_np, "relancé")
            st.session_state["followup_drafts"] = drafts
            st.success(f"✅ {len(drafts)} email(s) de relance générés.")
            st.rerun()

        # Affichage des drafts générés
        if st.session_state.get("followup_drafts"):
            for name, draft, _ in st.session_state["followup_drafts"]:
                st.markdown(f"**{name}**")
                st.code(draft, language=None)
            import io
            zip_content = "\n\n" + ("=" * 60 + "\n\n").join(
                f"{name}\n{draft}" for name, draft, _ in st.session_state["followup_drafts"]
            )
            st.download_button(
                "⬇️ Télécharger tous les emails de relance (.txt)",
                data=zip_content.encode("utf-8"),
                file_name=f"relances_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain",
                use_container_width=True,
            )

        # Liste individuelle avec bouton "A répondu"
        st.markdown("---")
        st.markdown("**Marquer comme répondu :**")
        for contact in due:
            col_name, col_btn = st.columns([4, 1])
            with col_name:
                date_str = contact.get("first_contact_date", "?")
                email_str = contact.get("email", "—")
                st.markdown(f"**{contact['name']}** — contacté le {date_str} — `{email_str}`")
            with col_btn:
                if st.button("✅ Répondu", key=f"responded_{contact['place_id']}"):
                    mark_as_responded(contact["place_id"])
                    if crm_type == "notion" and crm_key:
                        from history_manager import get_notion_page_id
                        from services.crm.notion import NotionExporter
                        _np = get_notion_page_id(contact["place_id"])
                        if _np:
                            NotionExporter(crm_key, crm_extra.get("database_id", "")).update_status(_np, "répondu")
                    st.rerun()
